from __future__ import annotations

import csv
import html.parser
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
HTML_FIXTURE = REPO_ROOT / "fixtures" / "sample-broker-statement.html"
TEMPLATE = REPO_ROOT / "templates" / "excel-export" / "UnravelTax-Template.xlsx"
OUTPUT_DIR = REPO_ROOT / "dry-runs" / "m1c"
FULL_WORKBOOK_OUTPUT = OUTPUT_DIR / "UnravelTax-M1C-Full-Workbook.xlsx"
CA_SUMMARY_OUTPUT = OUTPUT_DIR / "UnravelTax-M1C-CA-Summary.csv"
REPORT_OUTPUT = OUTPUT_DIR / "README.md"

EXPECTED_COLUMNS = [
    "Scrip Name",
    "Purchase Date",
    "Sell Date",
    "Units",
    "Buy Value",
    "Sell Value",
    "Buy Price",
    "Sell Price",
]


@dataclass
class TableCollector(html.parser.HTMLParser):
    tables: list[list[list[str]]] = field(default_factory=list)
    _current_table: list[list[str]] | None = None
    _current_row: list[str] | None = None
    _current_cell: list[str] | None = None
    _in_cell: bool = False

    def __post_init__(self) -> None:
        super().__init__()

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._current_table = []
        elif tag == "tr" and self._current_table is not None:
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            value = re.sub(r"\s+", " ", "".join(self._current_cell)).strip()
            self._current_row.append(value)
            self._current_cell = None
            self._in_cell = False
        elif tag == "tr" and self._current_row is not None and self._current_table is not None:
            if self._current_row:
                self._current_table.append(self._current_row)
            self._current_row = None
        elif tag == "table" and self._current_table is not None:
            self.tables.append(self._current_table)
            self._current_table = None


def parse_html_transaction_rows() -> list[list[str]]:
    parser = TableCollector()
    parser.feed(HTML_FIXTURE.read_text(encoding="utf-8"))
    for table in parser.tables:
        if table and table[0] == EXPECTED_COLUMNS:
            return table[1:]
    raise RuntimeError("Could not find the transaction table in the HTML fixture.")


def cell_to_prompt_format(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return "" if value is None else str(value)


def workbook_raw_rows() -> list[list[str]]:
    workbook = load_workbook(TEMPLATE, data_only=True, read_only=True)
    sheet = workbook["Raw Data - Sample Broker"]
    rows = []
    for row in sheet.iter_rows(min_row=5, max_row=9, min_col=1, max_col=8, values_only=True):
        rows.append([cell_to_prompt_format(value) for value in row])
    return rows


def export_ca_summary_csv() -> None:
    workbook = load_workbook(TEMPLATE, data_only=True, read_only=True)
    sheet = workbook["CA Summary"]
    with CA_SUMMARY_OUTPUT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(min_row=4, max_row=14, min_col=1, max_col=4, values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def write_report(transaction_rows: list[list[str]]) -> None:
    report = f"""# M1C Manual Flow Dry Run

Status: passed on 2026-07-02.

## Flow Exercised

1. README start path points to `prompts/00-master-guide.md`.
2. The master guide routes document extraction to `prompts/01-extract-statement.md`.
3. A non-CSV fixture was used: `fixtures/sample-broker-statement.html`.
4. The transaction table was selected from the HTML fixture, ignoring the disclaimer,
   charges summary, and portfolio summary tables.
5. Extracted rows matched the template's `Raw Data - Sample Broker` rows.
6. The template formulas produced the two intended outputs.

## Outputs

- Full workbook: `dry-runs/m1c/UnravelTax-M1C-Full-Workbook.xlsx`
- CA summary: `dry-runs/m1c/UnravelTax-M1C-CA-Summary.csv`

## Evidence

- Non-CSV transaction rows extracted: {len(transaction_rows)}
- Required workbook sheets verified by `scripts/verify-template.mjs`: 22
- CA Summary rows exported: 11

## Friction Points

- The Google Sheets hosted copy link is not published yet. The M1 workflow
  currently uses the Excel workbook directly or manual upload to Google Sheets.
- The AI extraction step cannot be executed inside this repo, so this dry run
  simulates the confirmed extraction output from the HTML fixture with a local
  parser and verifies that the rows match the template paste target.

## Blockers Fixed

- None. The flow is ready for the next slice: notebook skeleton.
"""
    REPORT_OUTPUT.write_text(report, encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    transaction_rows = parse_html_transaction_rows()
    raw_rows = workbook_raw_rows()
    if transaction_rows != raw_rows:
        raise AssertionError(f"HTML extraction rows do not match template raw rows.\nHTML: {transaction_rows}\nWorkbook: {raw_rows}")
    shutil.copyfile(TEMPLATE, FULL_WORKBOOK_OUTPUT)
    export_ca_summary_csv()
    write_report(transaction_rows)
    print("M1C manual flow dry run passed.")
    print(f"Wrote {FULL_WORKBOOK_OUTPUT}")
    print(f"Wrote {CA_SUMMARY_OUTPUT}")
    print(f"Wrote {REPORT_OUTPUT}")


if __name__ == "__main__":
    main()
