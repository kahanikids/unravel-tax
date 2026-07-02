from __future__ import annotations

import json
import csv
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
NOTEBOOK = REPO_ROOT / "notebooks" / "build-workbook.ipynb"
MANIFEST = REPO_ROOT / "notebooks" / "output" / "m2c-notebook-exports-manifest.json"
REFERENCE_CA_SUMMARY = REPO_ROOT / "dry-runs" / "m1c" / "UnravelTax-M1C-CA-Summary.csv"


def main() -> None:
    notebook = json.loads(NOTEBOOK.read_text(encoding="utf-8"))
    namespace = {"__name__": "__notebook_validation__"}

    for index, cell in enumerate(notebook["cells"], start=1):
        if cell.get("cell_type") != "code":
            continue
        source = "".join(cell.get("source", []))
        print(f"Running code cell {index}")
        exec(compile(source, f"{NOTEBOOK.name}:cell-{index}", "exec"), namespace)

    if not MANIFEST.exists():
        raise AssertionError(f"Notebook did not write manifest: {MANIFEST}")

    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    if manifest.get("milestone") != "M2C notebook exports":
        raise AssertionError(f"Unexpected manifest milestone: {manifest.get('milestone')}")
    outputs = manifest.get("outputs", [])
    if len(outputs) != 2:
        raise AssertionError("Notebook manifest should list the two generated outputs.")
    summary = manifest.get("selected_summary", {})
    expected = {
        "rows": 5,
        "intraday_gain": 800.0,
        "stcg": -500.0,
        "ltcg": 5500.0,
    }
    for key, value in expected.items():
        if summary.get(key) != value:
            raise AssertionError(f"Unexpected {key}: {summary.get(key)}")
    reference = manifest.get("reference_m1_summary", {})
    for key in ("intraday_gain", "stcg", "ltcg"):
        if reference.get(key) != summary.get(key):
            raise AssertionError(f"{key} does not match Milestone 1 reference: {reference.get(key)} vs {summary.get(key)}")
    if manifest.get("pdf_extracted_text_route") != "guided_prompt":
        raise AssertionError("PDF/free-form fixture should route to the guided prompt.")

    output_paths = [REPO_ROOT / output for output in outputs]
    for output_path in output_paths:
        if not output_path.exists():
            raise AssertionError(f"Missing notebook output: {output_path}")

    generated_ca_summary = next(path for path in output_paths if path.suffix == ".csv")
    generated_workbook = next(path for path in output_paths if path.suffix == ".xlsx")

    def normalize(value: str) -> str | int | float:
        if value == "":
            return value
        try:
            number = float(value)
        except ValueError:
            return value
        return int(number) if number.is_integer() else number

    def read_amounts(path: Path) -> dict[str, str | int | float]:
        with path.open(newline="", encoding="utf-8") as handle:
            return {row["Head"]: normalize(row["Amount"]) for row in csv.DictReader(handle)}

    if read_amounts(generated_ca_summary) != read_amounts(REFERENCE_CA_SUMMARY):
        raise AssertionError("Generated notebook CA Summary does not match M1C reference.")

    workbook = load_workbook(generated_workbook, data_only=True, read_only=True)
    required_sheets = {"CA Summary", "Transactions", "Detailed Summary", "Manifest"}
    if not required_sheets.issubset(set(workbook.sheetnames)):
        raise AssertionError(f"Generated workbook missing required sheets: {required_sheets - set(workbook.sheetnames)}")

    tx_sheet = workbook["Transactions"]
    if tx_sheet.max_row != 6:
        raise AssertionError(f"Expected 5 transaction rows plus header, found {tx_sheet.max_row}")

    print("Notebook export validation passed.")


if __name__ == "__main__":
    main()
