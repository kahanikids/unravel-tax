from __future__ import annotations

import csv
import html.parser
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


EXPECTED_COLUMNS = [
    "Scrip Name",
    "Purchase Date",
    "Sell Date",
    "Units",
    "Buy Value",
    "Sell Value",
    "Buy Price",
    "Sell Price",
]

SUPPLEMENTAL_INPUTS = {
    "dividends": 4000.0,
    "interest_other_income": 24000.0,
    "eligible_interest_deduction": 0.0,
    "deductible_transaction_charges": 160.0,
    "carry_forward_losses_available": 500.0,
}


@dataclass(frozen=True)
class Transaction:
    scrip_name: str
    purchase_date: datetime
    sell_date: datetime
    units: float
    buy_value: float
    sell_value: float
    buy_price: float
    sell_price: float
    hold_period_days: int
    tax_class: str
    gain_loss: float

    def comparable(self) -> tuple:
        return (
            self.scrip_name,
            self.purchase_date.strftime("%d-%b-%Y"),
            self.sell_date.strftime("%d-%b-%Y"),
            self.units,
            self.buy_value,
            self.sell_value,
            self.buy_price,
            self.sell_price,
            self.hold_period_days,
            self.tax_class,
            self.gain_loss,
        )


@dataclass
class TableCollector(html.parser.HTMLParser):
    tables: list[list[list[str]]] = field(default_factory=list)
    _current_table: list[list[str]] | None = None
    _current_row: list[str] | None = None
    _current_cell: list[str] | None = None
    _in_cell: bool = False

    def __post_init__(self) -> None:
        super().__init__()

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._current_table = []
        elif tag == "tr" and self._current_table is not None:
            self._current_row = []
        elif tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            value = re.sub(r"\s+", " ", "".join(self._current_cell)).strip()
            self._current_row.append(value)
            self._current_cell = None
            self._in_cell = False
        elif tag == "tr" and self._current_row is not None and self._current_table is not None:
            if self._current_row:
                self._current_table.append(self._current_row)
            self._current_row = None
        elif tag == "table" and self._current_table is not None:
            self.tables.append(self._current_table)
            self._current_table = None


def parse_date(value: str | datetime) -> datetime:
    if isinstance(value, datetime):
        return value
    return datetime.strptime(str(value), "%d-%b-%Y")


def parse_number(value: str | int | float) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").strip()
    return float(cleaned)


def normalize_rows(rows: Iterable[dict[str, object]]) -> list[Transaction]:
    transactions = []
    for row in rows:
        purchase_date = parse_date(row["Purchase Date"])
        sell_date = parse_date(row["Sell Date"])
        hold_period_days = (sell_date - purchase_date).days
        if hold_period_days == 0:
            tax_class = "Intraday"
        elif hold_period_days > 365:
            tax_class = "LT"
        else:
            tax_class = "ST"
        buy_value = parse_number(row["Buy Value"])
        sell_value = parse_number(row["Sell Value"])
        transactions.append(
            Transaction(
                scrip_name=str(row["Scrip Name"]).strip(),
                purchase_date=purchase_date,
                sell_date=sell_date,
                units=parse_number(row["Units"]),
                buy_value=buy_value,
                sell_value=sell_value,
                buy_price=parse_number(row["Buy Price"]),
                sell_price=parse_number(row["Sell Price"]),
                hold_period_days=hold_period_days,
                tax_class=tax_class,
                gain_loss=sell_value - buy_value,
            )
        )
    return transactions


def parse_csv_fixture(path: Path) -> list[Transaction]:
    with path.open(newline="", encoding="utf-8") as handle:
        return normalize_rows(csv.DictReader(handle))


def parse_structured_text_fixture(path: Path) -> list[Transaction]:
    with path.open(newline="", encoding="utf-8") as handle:
        return normalize_rows(csv.DictReader(handle, delimiter="\t"))


def parse_excel_fixture(path: Path) -> list[Transaction]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() for value in rows[0]]
    if headers != EXPECTED_COLUMNS:
        raise ValueError(f"Unexpected Excel headers: {headers}")
    records = [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]
    return normalize_rows(records)


def parse_html_fixture(path: Path) -> list[Transaction]:
    parser = TableCollector()
    parser.feed(path.read_text(encoding="utf-8"))
    for table in parser.tables:
        if table and table[0] == EXPECTED_COLUMNS:
            records = [dict(zip(EXPECTED_COLUMNS, row)) for row in table[1:]]
            return normalize_rows(records)
    raise ValueError("Could not find transaction table in HTML fixture.")


def route_pdf_or_freeform(path: Path) -> dict[str, str]:
    return {
        "path": str(path),
        "route": "guided_prompt",
        "prompt": "prompts/01-extract-statement.md",
        "reason": "PDF/free-form text table reconstruction stays in the AI-assisted prompt path.",
    }


def load_transactions(kind: str, path: Path) -> list[Transaction] | dict[str, str]:
    if kind == "csv":
        return parse_csv_fixture(path)
    if kind == "excel":
        return parse_excel_fixture(path)
    if kind == "html":
        return parse_html_fixture(path)
    if kind == "structured_text":
        return parse_structured_text_fixture(path)
    if kind == "pdf_extracted_text":
        return route_pdf_or_freeform(path)
    raise ValueError(f"Unsupported fixture kind: {kind}")


def summarize_transactions(transactions: list[Transaction]) -> dict[str, float | int]:
    return {
        "rows": len(transactions),
        "intraday_gain": sum(row.gain_loss for row in transactions if row.tax_class == "Intraday"),
        "stcg": sum(row.gain_loss for row in transactions if row.tax_class == "ST"),
        "ltcg": sum(row.gain_loss for row in transactions if row.tax_class == "LT"),
    }


def ca_summary_rows(transactions: list[Transaction]) -> list[list[object]]:
    summary = summarize_transactions(transactions)
    return [
        ["Head", "Rule/Section", "Amount", "Notes"],
        ["Speculative / Intraday income", "Business income", summary["intraday_gain"], "Notebook calculation"],
        ["Short-Term Capital Gains", "111A", summary["stcg"], "Notebook calculation"],
        ["Long-Term Capital Gains", "112A", summary["ltcg"], "Notebook calculation"],
        ["Dividends", "Schedule OS", SUPPLEMENTAL_INPUTS["dividends"], "Synthetic fixture supplemental input"],
        ["Interest & other income", "Schedule OS", SUPPLEMENTAL_INPUTS["interest_other_income"], "Synthetic fixture supplemental input"],
        ["Eligible interest deduction", "80TTA/80TTB", SUPPLEMENTAL_INPUTS["eligible_interest_deduction"], "Synthetic fixture supplemental input"],
        ["Deductible transaction charges", "Expense split", SUPPLEMENTAL_INPUTS["deductible_transaction_charges"], "Synthetic fixture supplemental input"],
        ["Carry-forward losses available", "CFL", SUPPLEMENTAL_INPUTS["carry_forward_losses_available"], "Synthetic fixture supplemental input"],
        ["Recommended ITR form", "", "ITR-3", "Business/speculative income is present"],
        ["CA review recommendation", "", "Get CA review before filing", "ITR-3 / speculative income trigger"],
    ]


def write_ca_summary_csv(transactions: list[Transaction], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows([[csv_value(value) for value in row] for row in ca_summary_rows(transactions)])


def csv_value(value: object) -> object:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F7A66")
    for cell in row:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")


def write_full_workbook(transactions: list[Transaction], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    ca_sheet = workbook.create_sheet("CA Summary")
    for row in ca_summary_rows(transactions):
        ca_sheet.append(row)
    _style_header(ca_sheet[1])

    tx_sheet = workbook.create_sheet("Transactions")
    tx_sheet.append([
        *EXPECTED_COLUMNS,
        "Hold Period (Days)",
        "Tax Class",
        "Gain/(Loss)",
    ])
    _style_header(tx_sheet[1])
    for tx in transactions:
        tx_sheet.append([
            tx.scrip_name,
            tx.purchase_date,
            tx.sell_date,
            tx.units,
            tx.buy_value,
            tx.sell_value,
            tx.buy_price,
            tx.sell_price,
            tx.hold_period_days,
            tx.tax_class,
            tx.gain_loss,
        ])

    detail_sheet = workbook.create_sheet("Detailed Summary")
    summary = summarize_transactions(transactions)
    detail_sheet.append(["Metric", "Value", "Notes"])
    _style_header(detail_sheet[1])
    detail_sheet.append(["Rows parsed", summary["rows"], "All lightweight fixture formats validate to this shape."])
    detail_sheet.append(["Intraday gain", summary["intraday_gain"], "Taxed as speculative/business income."])
    detail_sheet.append(["STCG", summary["stcg"], "Section 111A bucket."])
    detail_sheet.append(["LTCG", summary["ltcg"], "Section 112A bucket before exemption."])
    detail_sheet.append(["PDF/free-form route", "prompts/01-extract-statement.md", "No custom PDF parser in notebook."])

    manifest_sheet = workbook.create_sheet("Manifest")
    manifest_sheet.append(["Field", "Value"])
    _style_header(manifest_sheet[1])
    manifest_sheet.append(["Generated by", "notebooks/build-workbook.ipynb"])
    manifest_sheet.append(["Milestone", "M2C notebook exports"])
    manifest_sheet.append(["Source", "Synthetic fixtures only"])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 48)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "dd-mmm-yyyy"
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    workbook.save(output_path)


def read_ca_summary_csv(path: Path) -> dict[str, str]:
    with path.open(newline="", encoding="utf-8") as handle:
        return {row["Head"]: row["Amount"] for row in csv.DictReader(handle)}


def validate_ca_summary_against_reference(generated_path: Path, reference_path: Path) -> None:
    generated = read_ca_summary_csv(generated_path)
    reference = read_ca_summary_csv(reference_path)
    keys = [
        "Speculative / Intraday income",
        "Short-Term Capital Gains",
        "Long-Term Capital Gains",
        "Dividends",
        "Interest & other income",
        "Eligible interest deduction",
        "Deductible transaction charges",
        "Carry-forward losses available",
        "Recommended ITR form",
        "CA review recommendation",
    ]
    for key in keys:
        if normalize_csv_amount(generated.get(key)) != normalize_csv_amount(reference.get(key)):
            raise AssertionError(f"CA Summary mismatch for {key}: {generated.get(key)} != {reference.get(key)}")


def normalize_csv_amount(value: str | None) -> str | float | None:
    if value is None or value == "":
        return value
    try:
        number = float(value)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def validate_fixture_parity(fixture_options: dict[str, Path]) -> dict[str, object]:
    parsed = {
        kind: load_transactions(kind, path)
        for kind, path in fixture_options.items()
        if kind != "pdf_extracted_text"
    }
    baseline = parsed["csv"]
    assert isinstance(baseline, list)
    baseline_rows = [row.comparable() for row in baseline]
    for kind, rows in parsed.items():
        assert isinstance(rows, list)
        if [row.comparable() for row in rows] != baseline_rows:
            raise AssertionError(f"{kind} fixture does not match CSV baseline.")
    pdf_route = load_transactions("pdf_extracted_text", fixture_options["pdf_extracted_text"])
    assert isinstance(pdf_route, dict)
    return {
        "parsed_fixture_kinds": sorted(parsed),
        "pdf_extracted_text_route": pdf_route["route"],
        "summary": summarize_transactions(baseline),
    }
